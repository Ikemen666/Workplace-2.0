#!/usr/bin/env python3
"""
extract_xlsx.py — Characterize Excel/XLSX workbooks for ingestion.

Does NOT dump all rows. Produces structural metadata + a DeepSeek-powered
narrative summarizing what the workbook is about.

Usage:
    cd skills
    uv run --with openpyxl --with pandas --with openai --with requests \
        python extract-xlsx/extract_xlsx.py /path/to/file.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

BASE = Path(__file__).parent.parent
sys.path.insert(0, str(BASE / "_lib"))

SAMPLE_ROWS = 8        # rows to include in sample per sheet
FORMULA_SCAN_ROWS = 50 # rows to scan for notable formulas
NARRATE_THRESHOLD = 0  # always narrate (workbook characterization is the whole point)


def detect_language(text: str) -> str:
    if not text:
        return "en"
    cjk = sum(1 for c in text if "\u4e00" <= c <= "\u9fff")
    if cjk / max(len(text), 1) > 0.1:
        return "zh"
    return "en"


def _numeric_stats(values: list) -> dict | None:
    """Compute min/max/mean for a list of numeric values."""
    nums = [v for v in values if isinstance(v, (int, float)) and v is not None]
    if len(nums) < 2:
        return None
    return {
        "min": round(min(nums), 4),
        "max": round(max(nums), 4),
        "mean": round(sum(nums) / len(nums), 4),
        "count": len(nums),
    }


def analyze_sheet(ws) -> dict:
    """Extract structural metadata from a single worksheet."""
    import openpyxl  # type: ignore

    rows_iter = list(ws.iter_rows(values_only=False))
    row_count = len(rows_iter)
    col_count = ws.max_column or 0

    if row_count == 0:
        return {
            "name": ws.title,
            "row_count": 0,
            "col_count": 0,
            "columns": [],
            "numeric_stats": {},
            "sample_rows": [],
            "named_ranges": [],
            "notable_formulas": [],
        }

    # Headers: first non-empty row
    headers = []
    header_row_idx = 0
    for i, row in enumerate(rows_iter[:5]):
        vals = [cell.value for cell in row]
        if any(v is not None for v in vals):
            headers = [str(v) if v is not None else f"Col{j+1}" for j, v in enumerate(vals)]
            header_row_idx = i
            break

    data_rows = rows_iter[header_row_idx + 1:]
    data_row_count = len(data_rows)

    # Sample rows (values only)
    sample = []
    for row in data_rows[:SAMPLE_ROWS]:
        sample.append([cell.value for cell in row])

    # Numeric stats per column (scan all data rows)
    col_values: dict[int, list] = {j: [] for j in range(len(headers))}
    for row in data_rows:
        for j, cell in enumerate(row):
            if j < len(headers) and isinstance(cell.value, (int, float)):
                col_values[j].append(cell.value)

    numeric_stats = {}
    for j, col_name in enumerate(headers):
        stats = _numeric_stats(col_values.get(j, []))
        if stats:
            numeric_stats[col_name] = stats

    # Notable formulas (cells starting with "=")
    notable_formulas: list[str] = []
    seen_formulas: set[str] = set()
    for row in rows_iter[:FORMULA_SCAN_ROWS]:
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                # Normalize: strip cell references for dedup
                import re
                normalized = re.sub(r"[A-Z]+\d+", "REF", val)
                if normalized not in seen_formulas and len(notable_formulas) < 10:
                    notable_formulas.append(val)
                    seen_formulas.add(normalized)

    return {
        "name": ws.title,
        "row_count": data_row_count,
        "col_count": col_count,
        "columns": headers,
        "numeric_stats": numeric_stats,
        "sample_rows": sample,
        "named_ranges": [],  # populated at workbook level
        "notable_formulas": notable_formulas,
    }


def extract_xlsx(xlsx_path_str: str) -> dict:
    import openpyxl  # type: ignore

    xlsx_path = Path(xlsx_path_str).expanduser().resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # Named ranges at workbook level
    named_ranges = []
    try:
        for name in wb.defined_names:
            named_ranges.append(str(name))
    except Exception:
        pass

    sheets = []
    for ws in wb.worksheets:
        info = analyze_sheet(ws)
        info["named_ranges"] = [r for r in named_ranges if ws.title in r]
        sheets.append(info)

    # Structural summary
    total_rows = sum(s["row_count"] for s in sheets)
    sheet_names = [s["name"] for s in sheets]
    col_overview = []
    for s in sheets:
        if s["columns"]:
            col_overview.append(f"  {s['name']}: {', '.join(s['columns'][:8])}")

    structured_summary = (
        f"Workbook: {xlsx_path.name}\n"
        f"Sheets ({len(sheets)}): {', '.join(sheet_names)}\n"
        f"Total data rows: {total_rows:,}\n"
        f"Column overview:\n" + "\n".join(col_overview)
    )

    # Build narration prompt
    sheet_details = []
    for s in sheets:
        sample_str = ""
        if s["sample_rows"]:
            rows_str = "\n".join(
                "  " + " | ".join(str(v) for v in row[:8])
                for row in s["sample_rows"][:5]
            )
            sample_str = f"\nSample rows:\n{rows_str}"

        stats_str = ""
        if s["numeric_stats"]:
            stats_parts = [
                f"{col}: min={v['min']}, max={v['max']}, mean={v['mean']}"
                for col, v in list(s["numeric_stats"].items())[:5]
            ]
            stats_str = f"\nNumeric stats: {'; '.join(stats_parts)}"

        formula_str = ""
        if s["notable_formulas"]:
            formula_str = f"\nFormulas: {', '.join(s['notable_formulas'][:5])}"

        sheet_details.append(
            f"## Sheet: {s['name']} ({s['row_count']:,} data rows, {s['col_count']} columns)\n"
            f"Columns: {', '.join(s['columns'][:12])}"
            f"{sample_str}{stats_str}{formula_str}"
        )

    narrate_prompt = (
        f"You are analyzing an Excel workbook. Describe what this workbook is for, "
        f"what kind of data it tracks, and any notable patterns or purposes you can infer. "
        f"Be specific and factual. 2-4 sentences.\n\n"
        f"File: {xlsx_path.name}\n\n"
        + "\n\n".join(sheet_details)
    )

    # Call DeepSeek for narration
    from llm_router import call_model
    resp = call_model("xlsx_narrate", narrate_prompt, max_tokens=400)
    narrative_summary = resp.get("text", "").strip()

    # Language detection
    all_text = " ".join(
        str(v)
        for s in sheets
        for row in s["sample_rows"]
        for v in row
        if isinstance(v, str)
    )
    language_hint = detect_language(all_text + narrative_summary)

    author = ""
    try:
        props = wb.properties
        author = getattr(props, "creator", "") or ""
    except Exception:
        pass

    # raw_text for distiller compatibility (narrative + structure)
    raw_text = f"{narrative_summary}\n\n{structured_summary}"

    return {
        "raw_text": raw_text,
        "sheets": sheets,
        "structured_summary": structured_summary,
        "narrative_summary": narrative_summary,
        "metadata": {
            "filename": xlsx_path.name,
            "sheet_count": len(sheets),
            "total_rows": total_rows,
        },
        "language_hint": language_hint,
        "images_dir": "",
        "source_metadata": {
            "source_type": "xlsx",
            "original_url_or_path": str(xlsx_path),
            "filename": xlsx_path.name,
            "sheet_count": len(sheets),
            "author": author,
            "language": language_hint,
        },
    }


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Extract content from XLSX file")
    parser.add_argument("xlsx", help="Path to XLSX file")
    args = parser.parse_args()

    result = extract_xlsx(args.xlsx)
    preview = dict(result)
    preview.pop("sheets", None)  # verbose — show separately
    print(json.dumps(preview, indent=2, ensure_ascii=False))
    print(f"\nSheets ({len(result['sheets'])}):")
    for s in result["sheets"]:
        print(f"  {s['name']}: {s['row_count']:,} rows, columns: {', '.join(s['columns'][:6])}")


if __name__ == "__main__":
    main()
